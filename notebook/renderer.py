from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def render():
    wb = load_workbook('assets/template.xlsx')
    sheet = wb['Dokument']

    # Die eigentliche Logik zur Erstellung.
    start_column = 2
    start_row = 3

    for i in range(start_column, 10):
        for j in range(start_row, 15):
            sheet[get_column_letter(i) + str(j)] = 1

    wb.save('assets/result.xlsx')
    